import re
import threading
import os
import tkinter as tk
from tkinter import filedialog

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ═══════════════════════════════════════════════════════════════
#  PALETA  UI
# ═══════════════════════════════════════════════════════════════
BG        = "#0F1923"
CARD      = "#1A2535"
ACCENT    = "#25D366"
ACCENT2   = "#128C7E"
TEXT      = "#FFFFFF"
SUBTEXT   = "#8A9BB0"
ERROR_CLR = "#FF5C5C"

FONT_TITLE = ("Segoe UI", 18, "bold")
FONT_BODY  = ("Segoe UI", 10)
FONT_SMALL = ("Segoe UI", 9)
FONT_BOLD  = ("Segoe UI", 11, "bold")
FONT_MONO  = ("Consolas", 9)

# ═══════════════════════════════════════════════════════════════
#  PARSER
# ═══════════════════════════════════════════════════════════════
def parse_whatsapp_chat(filepath):
    padrao = re.compile(
        r'(\d{1,2}/\d{1,2}/\d{2,4}),?\s(\d{1,2}:\d{2})(?:\s?[APap][Mm])?\s?-\s([^:]+):\s(.+)'
    )
    mensagens, atual = [], None
    with open(filepath, 'r', encoding='utf-8') as f:
        for linha in f:
            linha = linha.strip()
            m = padrao.match(linha)
            if m:
                if atual:
                    mensagens.append(atual)
                data, hora, autor, texto = m.groups()
                atual = {'data': data, 'hora': hora,
                         'autor': autor.strip(), 'mensagem': texto.strip()}
            elif atual and linha:
                atual['mensagem'] += ' ' + linha
    if atual:
        mensagens.append(atual)
    return pd.DataFrame(mensagens)

# ═══════════════════════════════════════════════════════════════
#  GERADOR EXCEL
# ═══════════════════════════════════════════════════════════════
# Cores Excel (hex sem #)
XL_GREEN_DARK  = "1A4731"
XL_GREEN       = "25D366"
XL_GREEN_LIGHT = "D6F5E3"
XL_TEAL        = "128C7E"
XL_HEADER_BG   = "075E54"   # verde escuro WhatsApp
XL_WHITE       = "FFFFFF"
XL_GRAY_LIGHT  = "F5F7FA"
XL_GRAY        = "E2E8F0"
XL_BLUE_LIGHT  = "EBF4FF"
XL_PURPLE_LIGHT= "F3EEFF"

# Paleta de cores por autor (rotativa)
AUTHOR_COLORS = [
    ("DCF8C6", "1A4731"),  # verde claro / texto escuro
    ("E8F4FD", "1A3A5C"),  # azul claro
    ("FFF3CD", "5C4B00"),  # amarelo
    ("F8D7DA", "5C1A1A"),  # vermelho claro
    ("E2D9F3", "3A1A5C"),  # roxo claro
    ("D1ECF1", "0C4A5C"),  # ciano claro
]

def thin_border():
    s = Side(border_style="thin", color="D0D7DE")
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color="25D366"):
    b = Side(border_style="medium", color=color)
    return Border(bottom=b)

def make_excel(df: pd.DataFrame, output_path: str):
    wb = Workbook()

    # ── ABA 1: Conversa ──────────────────────────────────────
    ws = wb.active
    ws.title = "💬 Conversa"
    ws.sheet_view.showGridLines = False

    # Mapa de cor por autor
    autores = df['autor'].unique().tolist()
    cor_autor = {a: AUTHOR_COLORS[i % len(AUTHOR_COLORS)] for i, a in enumerate(autores)}

    # ── Cabeçalho principal ──────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = "💬  WhatsApp — Conversa Exportada"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color=XL_WHITE)
    ws["A1"].fill      = PatternFill("solid", fgColor=XL_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Cabeçalhos de coluna ─────────────────────────────────
    headers = ["#", "Data", "Hora", "Participante", "Mensagem"]
    col_widths = [6, 14, 10, 26, 80]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=XL_WHITE)
        cell.fill      = PatternFill("solid", fgColor=XL_TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    # ── Dados ────────────────────────────────────────────────
    for i, row in df.iterrows():
        r = i + 3  # linha Excel (começa em 3)
        autor  = row['autor']
        bg, fg = cor_autor.get(autor, ("FFFFFF", "000000"))
        fill   = PatternFill("solid", fgColor=bg)
        fonte  = Font(name="Arial", size=9, color=fg)
        fonte_num = Font(name="Arial", size=9, color="888888")

        values = [i + 1, row['data'], row['hora'], autor, row['mensagem']]
        aligns = ["center", "center", "center", "left", "left"]

        for col, (val, align) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill      = fill if col > 1 else PatternFill("solid", fgColor=XL_GRAY_LIGHT)
            cell.font      = fonte_num if col == 1 else fonte
            cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=(col == 5))
            cell.border    = thin_border()

        ws.row_dimensions[r].height = 15

    # Congela cabeçalho
    ws.freeze_panes = "A3"

    # Filtro automático
    ws.auto_filter.ref = f"A2:E{len(df)+2}"

    # ── ABA 2: Resumo ────────────────────────────────────────
    ws2 = wb.create_sheet("📊 Resumo")
    ws2.sheet_view.showGridLines = False

    # Título
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "📊  Resumo da Conversa"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=14, color=XL_WHITE)
    ws2["A1"].fill      = PatternFill("solid", fgColor=XL_HEADER_BG)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 34

    # Cards de métricas
    metricas = [
        ("💬 Total de Mensagens",  len(df)),
        ("👥 Participantes",        len(autores)),
        ("📅 Data Inicial",         df['data'].iloc[0]),
        ("📅 Data Final",           df['data'].iloc[-1]),
    ]
    for col, (label, val) in enumerate(metricas, 1):
        ws2.column_dimensions[get_column_letter(col)].width = 24
        ws2.cell(row=2, column=col, value=label).font = Font(name="Arial", bold=True, size=9, color="555555")
        ws2.cell(row=2, column=col).fill = PatternFill("solid", fgColor=XL_GRAY)
        ws2.cell(row=2, column=col).alignment = Alignment(horizontal="center")
        c = ws2.cell(row=3, column=col, value=val)
        c.font      = Font(name="Arial", bold=True, size=14, color=XL_HEADER_BG)
        c.fill      = PatternFill("solid", fgColor=XL_GREEN_LIGHT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bottom_border()
        ws2.row_dimensions[3].height = 30

    # Cabeçalho tabela de participantes
    ws2.row_dimensions[5].height = 20
    for col, h in enumerate(["Participante", "Mensagens", "% do Total", "Primeira Msg", "Última Msg"], 1):
        c = ws2.cell(row=5, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=10, color=XL_WHITE)
        c.fill      = PatternFill("solid", fgColor=XL_TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 16

    # Dados por autor
    for i, autor in enumerate(autores):
        r   = i + 6
        sub = df[df['autor'] == autor]
        bg, fg = cor_autor[autor]
        fill = PatternFill("solid", fgColor=bg)
        font = Font(name="Arial", size=10, color=fg)
        total_row = len(df)

        vals = [
            autor,
            len(sub),
            f'=B{r}/B${6+len(autores)-1+1}',   # calculado abaixo via fórmula
            sub['data'].iloc[0],
            sub['data'].iloc[-1],
        ]
        # % com fórmula Excel
        ws2.cell(row=r, column=1, value=autor).font = font
        ws2.cell(row=r, column=1).fill      = fill
        ws2.cell(row=r, column=1).border    = thin_border()
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)

        ws2.cell(row=r, column=2, value=len(sub)).font = font
        ws2.cell(row=r, column=2).fill      = fill
        ws2.cell(row=r, column=2).border    = thin_border()
        ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center")

        pct_cell = ws2.cell(row=r, column=3, value=f"=B{r}/SUM(B6:B{5+len(autores)})")
        pct_cell.font        = font
        pct_cell.fill        = fill
        pct_cell.border      = thin_border()
        pct_cell.alignment   = Alignment(horizontal="center")
        pct_cell.number_format = "0.0%"

        ws2.cell(row=r, column=4, value=sub['data'].iloc[0]).font   = font
        ws2.cell(row=r, column=4).fill   = fill; ws2.cell(row=r, column=4).border = thin_border()
        ws2.cell(row=r, column=4).alignment = Alignment(horizontal="center")

        ws2.cell(row=r, column=5, value=sub['data'].iloc[-1]).font  = font
        ws2.cell(row=r, column=5).fill   = fill; ws2.cell(row=r, column=5).border = thin_border()
        ws2.cell(row=r, column=5).alignment = Alignment(horizontal="center")

    # Linha de total
    total_r = 6 + len(autores)
    ws2.row_dimensions[total_r].height = 18
    for col in range(1, 6):
        c = ws2.cell(row=total_r, column=col)
        c.fill   = PatternFill("solid", fgColor=XL_HEADER_BG)
        c.font   = Font(name="Arial", bold=True, size=10, color=XL_WHITE)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center")
    ws2.cell(row=total_r, column=1, value="TOTAL").alignment = Alignment(horizontal="left", indent=1)
    ws2.cell(row=total_r, column=2, value=f"=SUM(B6:B{total_r-1})")
    ws2.cell(row=total_r, column=3, value="100%")
    ws2.cell(row=total_r, column=3).number_format = "0.0%"

    # ── ABA 3: Por Dia ───────────────────────────────────────
    ws3 = wb.create_sheet("📅 Por Dia")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "📅  Mensagens por Dia"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=14, color=XL_WHITE)
    ws3["A1"].fill      = PatternFill("solid", fgColor=XL_HEADER_BG)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 34

    por_dia = df.groupby('data').size().reset_index(name='total')
    por_dia.columns = ['Data', 'Total de Mensagens']

    for col, h in enumerate(["Data", "Total de Mensagens"], 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=10, color=XL_WHITE)
        c.fill      = PatternFill("solid", fgColor=XL_TEAL)
        c.alignment = Alignment(horizontal="center")
        c.border    = thin_border()

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 22

    for i, row in por_dia.iterrows():
        r   = i + 3
        bg  = XL_GRAY_LIGHT if i % 2 == 0 else XL_WHITE
        fill = PatternFill("solid", fgColor=bg)
        for col, val in enumerate([row['Data'], row['Total de Mensagens']], 1):
            c = ws3.cell(row=r, column=col, value=val)
            c.fill      = fill
            c.font      = Font(name="Arial", size=10)
            c.border    = thin_border()
            c.alignment = Alignment(horizontal="center")

    # Total
    last_r = len(por_dia) + 3
    ws3.cell(row=last_r, column=1, value="TOTAL").font   = Font(name="Arial", bold=True, color=XL_WHITE)
    ws3.cell(row=last_r, column=1).fill   = PatternFill("solid", fgColor=XL_HEADER_BG)
    ws3.cell(row=last_r, column=1).border = thin_border()
    ws3.cell(row=last_r, column=1).alignment = Alignment(horizontal="center")
    ws3.cell(row=last_r, column=2, value=f"=SUM(B3:B{last_r-1})").font = Font(name="Arial", bold=True, color=XL_WHITE)
    ws3.cell(row=last_r, column=2).fill   = PatternFill("solid", fgColor=XL_HEADER_BG)
    ws3.cell(row=last_r, column=2).border = thin_border()
    ws3.cell(row=last_r, column=2).alignment = Alignment(horizontal="center")

    ws3.freeze_panes = "A3"

    wb.save(output_path)

# ═══════════════════════════════════════════════════════════════
#  BOTÃO ARREDONDADO
# ═══════════════════════════════════════════════════════════════
class RoundedButton(tk.Canvas):
    def __init__(self, parent, text, command,
                 btn_width=200, btn_height=44,
                 color=ACCENT, hover_color=ACCENT2,
                 fg=BG, font=FONT_BOLD, radius=22):
        super().__init__(parent, highlightthickness=0, cursor="hand2", bg=parent["bg"])
        self.configure(width=btn_width, height=btn_height)
        self._color = color; self._hover_color = hover_color
        self._fg = fg; self._font = font; self._radius = radius
        self._bw = btn_width; self._bh = btn_height
        self._text = text; self._command = command
        self._draw(self._color)
        self.bind("<Enter>",    lambda e: self._draw(self._hover_color))
        self.bind("<Leave>",    lambda e: self._draw(self._color))
        self.bind("<Button-1>", lambda e: self._command())

    def _draw(self, color):
        self.delete("all")
        r, w, h = self._radius, self._bw, self._bh
        self.create_arc(0,     0,     2*r, 2*r, start=90,  extent=90,  fill=color, outline=color)
        self.create_arc(w-2*r, 0,     w,   2*r, start=0,   extent=90,  fill=color, outline=color)
        self.create_arc(0,     h-2*r, 2*r, h,   start=180, extent=90,  fill=color, outline=color)
        self.create_arc(w-2*r, h-2*r, w,   h,   start=270, extent=90,  fill=color, outline=color)
        self.create_rectangle(r, 0,   w-r, h,   fill=color, outline=color)
        self.create_rectangle(0, r,   w,   h-r, fill=color, outline=color)
        self.create_text(w // 2, h // 2, text=self._text, fill=self._fg, font=self._font)

# ═══════════════════════════════════════════════════════════════
#  LINHA STAT
# ═══════════════════════════════════════════════════════════════
def stat_row(parent, icon, label, value, value_color=TEXT):
    row = tk.Frame(parent, bg=CARD)
    row.pack(fill="x", padx=16, pady=3)
    tk.Label(row, text=icon,  font=FONT_BODY, bg=CARD, fg=ACCENT, width=2, anchor="w").pack(side="left")
    tk.Label(row, text=label, font=FONT_BOLD, bg=CARD, fg=SUBTEXT                    ).pack(side="left")
    tk.Label(row, text=value, font=FONT_BODY, bg=CARD, fg=value_color,
             wraplength=330, justify="left").pack(side="left", padx=(6, 0))

# ═══════════════════════════════════════════════════════════════
#  APP
# ═══════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("WhatsApp Parser")
        self.resizable(False, False)
        self.configure(bg=BG)
        self._center(540, 340)
        self._arquivo = None
        self._anim_id = None
        self._build()

    def _center(self, w, h):
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=CARD, height=65)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="💬", font=("Segoe UI Emoji", 20), bg=CARD, fg=ACCENT).pack(side="left", padx=(18,6), pady=10)
        tk.Label(hdr, text="WhatsApp Parser", font=FONT_TITLE, bg=CARD, fg=TEXT  ).pack(side="left", pady=10)
        tk.Label(hdr, text="v3.0 · Excel", font=FONT_SMALL, bg=CARD, fg=SUBTEXT  ).pack(side="left", pady=(18,0))

        # Body
        self._body = tk.Frame(self, bg=BG)
        self._body.pack(fill="both", expand=True, padx=28, pady=18)

        # Card seleção
        zone = tk.Frame(self._body, bg=CARD)
        zone.pack(fill="x", pady=(0,12))
        tk.Label(zone, text="Arquivo exportado do WhatsApp", font=FONT_BOLD, bg=CARD, fg=TEXT).pack(anchor="w", padx=16, pady=(12,4))
        row = tk.Frame(zone, bg=CARD)
        row.pack(fill="x", padx=16, pady=(0,12))
        self._lbl_arquivo = tk.Label(row, text="Nenhum arquivo selecionado",
                                     font=FONT_BODY, bg="#111D2B", fg=SUBTEXT, anchor="w", padx=10, pady=7)
        self._lbl_arquivo.pack(side="left", fill="x", expand=True)
        RoundedButton(row, "Procurar", self._selecionar,
                      btn_width=100, btn_height=34,
                      color=ACCENT2, hover_color="#0e6b62", fg=TEXT, radius=17
                      ).pack(side="left", padx=(10,0))

        # Dica
        tk.Label(self._body,
                 text="ℹ  WhatsApp → abra a conversa → ⋮ → Mais → Exportar conversa → Sem mídia",
                 font=FONT_SMALL, bg=BG, fg=SUBTEXT, wraplength=480, justify="left"
                 ).pack(anchor="w", pady=(0,12))

        # Badge Excel
        badge = tk.Frame(self._body, bg="#1D6F42")
        badge.pack(anchor="w", pady=(0,12))
        tk.Label(badge, text="  📗 Exporta para Excel (.xlsx) com 3 abas formatadas  ",
                 font=FONT_SMALL, bg="#1D6F42", fg="#FFFFFF").pack(padx=4, pady=4)

        # Progresso
        self._prog = tk.Canvas(self._body, height=5, bg="#1A2535", highlightthickness=0)
        self._prog.pack(fill="x", pady=(0,4))

        # Status
        self._status_lbl = tk.Label(self._body, text="", font=FONT_SMALL, bg=BG, fg=SUBTEXT)
        self._status_lbl.pack(anchor="w", pady=(0,8))

        # Card resultado
        self._result_card = tk.Frame(self._body, bg=CARD)

        # Botão
        self._btn = RoundedButton(self._body, "Processar e Exportar para Excel", self._processar,
                                  btn_width=484, btn_height=44,
                                  color=ACCENT, hover_color=ACCENT2, fg=BG, radius=22)
        self._btn.pack(side="bottom")

    # ── Handlers ─────────────────────────────────────────────
    def _selecionar(self):
        path = filedialog.askopenfilename(
            title="Selecione o arquivo do WhatsApp",
            filetypes=[("Arquivos de texto", "*.txt"), ("Todos", "*.*")]
        )
        if path:
            self._arquivo = path
            self._lbl_arquivo.config(text=os.path.basename(path), fg=TEXT)
            self._set_status("")
            self._esconder_resultado()

    def _processar(self):
        if not self._arquivo:
            self._set_status("⚠  Selecione um arquivo primeiro.", ERROR_CLR)
            return
        self._esconder_resultado()
        self._set_status("⏳  Processando...", SUBTEXT)
        self._animar(0)
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        try:
            df = parse_whatsapp_chat(self._arquivo)
            if df.empty:
                self.after(0, self._parar_anim)
                self.after(0, lambda: self._set_status("❌  Nenhuma mensagem encontrada.", ERROR_CLR))
                return
            total   = len(df)
            autores = df['autor'].unique().tolist()
            inicio  = df['data'].iloc[0]
            fim     = df['data'].iloc[-1]
            self.after(0, lambda: self._pedir_destino(df, total, autores, inicio, fim))
        except Exception as ex:
            self.after(0, self._parar_anim)
            self.after(0, lambda: self._set_status(f"❌  Erro: {ex}", ERROR_CLR))

    def _pedir_destino(self, df, total, autores, inicio, fim):
        self._parar_anim()
        saida = filedialog.asksaveasfilename(
            title="Onde salvar o arquivo Excel?",
            initialfile="conversa_whatsapp.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if not saida:
            self._set_status("⚠  Salvamento cancelado.", SUBTEXT)
            return
        self._set_status("⏳  Gerando Excel formatado...", SUBTEXT)
        self._animar(0)
        threading.Thread(target=lambda: self._salvar(df, saida, total, autores, inicio, fim), daemon=True).start()

    def _salvar(self, df, saida, total, autores, inicio, fim):
        try:
            make_excel(df, saida)
            self.after(0, self._parar_anim)
            self.after(0, lambda: self._exibir_resultado(total, autores, inicio, fim, saida))
        except Exception as ex:
            self.after(0, self._parar_anim)
            self.after(0, lambda: self._set_status(f"❌  Erro ao salvar: {ex}", ERROR_CLR))

    def _exibir_resultado(self, total, autores, inicio, fim, saida):
        w = self._prog.winfo_width() or 484
        self._prog.delete("all")
        self._prog.create_rectangle(0, 0, w, 5, fill=ACCENT, outline=ACCENT)
        self._set_status("✅  Excel gerado com sucesso!", ACCENT)

        for widget in self._result_card.winfo_children():
            widget.destroy()

        tk.Label(self._result_card, text="Resultado", font=FONT_BOLD, bg=CARD, fg=SUBTEXT
                 ).pack(anchor="w", padx=16, pady=(10,6))
        stat_row(self._result_card, "📊", "Mensagens:",    str(total), TEXT)
        stat_row(self._result_card, "📅", "Período:",      f"{inicio}  →  {fim}", TEXT)
        stat_row(self._result_card, "👥", "Participantes:", str(len(autores)), TEXT)

        pframe = tk.Frame(self._result_card, bg="#111D2B")
        pframe.pack(fill="x", padx=16, pady=(4,8))
        for a in autores:
            tk.Label(pframe, text=f"  {a}", font=FONT_MONO, bg="#111D2B",
                     fg=SUBTEXT, anchor="w").pack(fill="x", padx=4, pady=1)

        tk.Label(self._result_card, text="📗  3 abas geradas: 💬 Conversa · 📊 Resumo · 📅 Por Dia",
                 font=FONT_SMALL, bg=CARD, fg=ACCENT).pack(anchor="w", padx=16, pady=(0,4))
        stat_row(self._result_card, "💾", "Salvo em:", os.path.basename(saida), ACCENT)
        tk.Frame(self._result_card, bg=CARD, height=10).pack()

        self._result_card.pack(fill="x", pady=(0,10), before=self._btn)
        self.update_idletasks()
        self.geometry(f"540x{self.winfo_reqheight()}+{self.winfo_x()}+{self.winfo_y()}")

    def _esconder_resultado(self):
        self._result_card.pack_forget()
        self.update_idletasks()
        self.geometry(f"540x340+{self.winfo_x()}+{self.winfo_y()}")

    def _set_status(self, msg, color=SUBTEXT):
        self._status_lbl.config(text=msg, fg=color)

    def _animar(self, pos):
        self._prog.delete("all")
        w  = self._prog.winfo_width() or 484
        bw = 120
        x  = pos % (w + bw) - bw
        self._prog.create_rectangle(x, 0, x+bw, 5, fill=ACCENT, outline=ACCENT)
        self._anim_id = self.after(16, lambda: self._animar(pos+8))

    def _parar_anim(self):
        if self._anim_id:
            self.after_cancel(self._anim_id)
            self._anim_id = None


if __name__ == "__main__":
    app = App()
    app.mainloop()